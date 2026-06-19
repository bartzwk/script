import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo

# =========================================================
# PROJECT CONFIGURATIONS
# Uso:  python create_assembled_excel.py          -> ambos
#       python create_assembled_excel.py CP1      -> solo CP1
#       python create_assembled_excel.py FA1      -> solo FA1
#
# Ajusta la subcarpeta en root_dir si la estructura del
# disco difiere (p.ej. sin "1. Dokumentacja projektowa").
# =========================================================

DEBUG = False  # Cambia a True para ver cabeceras raw y reconstrucciones

# Generated files go to ./out next to this script
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "out")

PROJECT_CONFIGS = {
    "CP1": {
        "root_dir":       r"P:\Aleksander Popławski 2\121-CP1\1. Dokumentacja projektowa",
        "output_name":    "master CP1 assembled.xlsx",
        "search_folders": ["A", "A2", "B", "C"],
        "prefix_pattern": r"^CP\d+[-_ ]*",
    },
    "FA1": {
        "root_dir":       r"P:\Aleksander Popławski 2\131-FA1\1. Dokumentacja projektowa",
        "output_name":    "master FA1 assembled.xlsx",
        "search_folders": ["A1", "A2", "B", "C1", "C2"],
        "prefix_pattern": r"^FA\d*[-_ ]*",
    },
}

OUTPUT_COLUMNS = [
    "Name",
    "Zone",
    "Symbol",
    "Element",
    "Element number",
    "REV",
    "Construction",
    "Steel type",
    "Quantity",
    "Description",
    "Length (mm)",
    "Weight netto (kg/pcs.)",
    "Weight netto (kg)",
    "Weight brutto (kg)",
    "Weld allowance (%)",
    "Weight netto (kg/pcs.) source",
    "Weight netto (kg) source",
    "Weight brutto (kg) source",
    "Netto check",
    "Surface area (m2/pcs.)",
    "Number items included",
    "All items included",
    "Source file path",
    "Source row number",
]

# =========================================================
# TEXT HELPERS
# =========================================================

def safe_int(value: object) -> Optional[int]:
    f = safe_float(value)
    if f is None:
        return None
    if abs(f - round(f)) < 1e-9:
        return int(round(f))
    return None


def excel_number(value: object) -> object:
    f = safe_float(value)
    if f is None:
        return clean_value(value)
    if abs(f - round(f)) < 1e-9:
        return int(round(f))
    return f


def extract_revision_from_text(text: object) -> Optional[int]:
    s = str(text or "")
    m = re.search(r"(?:^|[^A-Z0-9])R\s*(\d+)(?:[^A-Z0-9]|$)", s, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"Revision\s*(\d+)", s, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def strip_accents(text: str) -> str:
    s = str(text).replace("ł", "l").replace("Ł", "L")
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", s)
        if not unicodedata.combining(ch)
    )


def normalize_text(text: object) -> str:
    if text is None:
        return ""
    s = str(text).replace("\xa0", " ").replace("\n", " ").replace("\r", " ")
    s = strip_accents(s).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_key(text: object) -> str:
    s = normalize_text(text)
    return s.replace(" ", "")


def clean_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return value
    return str(value).replace("\xa0", " ").strip()


def safe_float(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "")
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# =========================================================
# FILE CLASSIFICATION
# =========================================================

def get_top_area(rel_path: str) -> str:
    parts = Path(rel_path).parts
    return parts[0] if parts else ""


def is_summary_company_folder(path_parts_lower: List[str]) -> bool:
    for p in path_parts_lower:
        p_norm = normalize_text(p)
        if p_norm in {"kg+gt", "kg i gt section", "kg i gt", "kg&gt", "gt+kg"}:
            return True
    return False


def detect_company(path_parts: Tuple[str, ...], file_name: str) -> str:
    candidates = [file_name, *path_parts]
    joined = " | ".join(candidates)
    n = normalize_text(joined)
    if re.search(r"(^|[^a-z])gt([^a-z]|$)", n) and not re.search(r"kg i gt|kg\+gt|gt\+kg", n):
        if re.search(r"(^|[^a-z])kg([^a-z]|$)", n):
            return "MULTI"
        return "GT"
    if re.search(r"(^|[^a-z])kg([^a-z]|$)", n):
        return "KG"
    return "UNKNOWN"


def infer_doc_type(file_name: str) -> Optional[str]:
    n = normalize_text(file_name)
    if "zestawienie lacznikow" in n or "lista lacznikow" in n or "list of connectors" in n:
        return None
    if "zestawienie elementow wysylkowych" in n or "lista elementow wysylkowych" in n or "shipping items" in n:
        return "shipping"
    if "zestawienie stali profilowej" in n or "lista stali profilowej" in n or "list of profiles" in n:
        return "profiles"
    if "zestawienie blachy-obejmy" in n or "lista blachy-obejmy" in n or "steel sheets-clamp" in n:
        return "clamp_sheets"
    if "zestawienie c-channel" in n or "lista c-channel" in n or "list of c-channel" in n:
        return "c_channel"
    if "zestawienie blach" in n or "lista blach" in n or "list of sheets" in n:
        return "sheets"
    return None


def should_skip_file(full_path: str, cfg: Dict) -> Tuple[bool, str]:
    root_dir = cfg["root_dir"]
    search_folders = cfg["search_folders"]

    rel_path = os.path.relpath(full_path, root_dir)
    parts = Path(rel_path).parts
    parts_lower = [p.lower() for p in parts]
    file_name = os.path.basename(full_path)
    file_name_norm = normalize_text(file_name)

    if len(parts) < 2:
        return True, "not_inside_section_subfolder"

    if parts[0].upper() not in [f.upper() for f in search_folders]:
        return True, "outside_sections"

    if file_name.startswith("~$"):
        return True, "temp_file"

    if not file_name.lower().endswith(".xls"):
        return True, "not_xls"

    if is_summary_company_folder(parts_lower):
        return True, "summary_company_folder"

    if file_name_norm.startswith(normalize_text("master ")):
        return True, "generated_output"

    doc_type = infer_doc_type(file_name)
    if doc_type is None:
        return True, "unsupported_or_connectors"

    return False, doc_type


# =========================================================
# EXCEL READING
# =========================================================

def read_sheet_matrix(xls_path: str) -> List[List[object]]:
    book = xlrd.open_workbook(xls_path)
    sheet = book.sheet_by_index(0)
    return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]


def score_header_row(row: List[object]) -> int:
    hits = 0
    for cell in row:
        t = normalize_text(cell)
        if "numer" in t or "number" in t:
            hits += 1
        if "nazwa" in t or "name" in t:
            hits += 1
        if "ilosc" in t or "quantity" in t:
            hits += 1
        if "dlugosc" in t or "lenght" in t or "length" in t:
            hits += 1
        if "waga elementu" in t or "item weight" in t:
            hits += 1
        if "lacznie waga" in t or "total weight" in t:
            hits += 1
        if "wszystkie dolaczone pozycje" in t or "all items included" in t:
            hits += 2
    return hits


def find_header_row(matrix: List[List[object]]) -> Optional[int]:
    best_idx, best_score = None, -1
    for i, row in enumerate(matrix[:20]):
        score = score_header_row(row)
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score >= 2 else None


def build_header_map(header_row: List[object]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        t = normalize_text(cell)
        if not t:
            continue
        if "numer" in t or "number" in t:
            header_map.setdefault("number", idx)
        if "nazwa" in t or ("name" in t and "project" not in t):
            header_map.setdefault("name", idx)
        if "ilosc" in t or "quantity" in t:
            header_map.setdefault("quantity", idx)
        if "dlugosc" in t or "lenght" in t or "length" in t:
            header_map.setdefault("length", idx)
        if "szerokosc" in t or "width" in t:
            header_map.setdefault("width", idx)
        if "grubosc" in t or "thickness" in t:
            header_map.setdefault("thickness", idx)
        if "waga elementu" in t or "item weight" in t:
            header_map.setdefault("item_weight", idx)
        if "lacznie waga" in t or "total weight" in t:
            header_map.setdefault("total_weight", idx)
        if "powierzchnia elementu" in t or "surface area" in t:
            header_map.setdefault("surface", idx)
        if "klasa" in t or "class" in t:
            header_map.setdefault("class", idx)
        if "grade" in t:
            header_map.setdefault("grade", idx)
        if "waga na metr" in t or "weight per meter" in t:
            header_map.setdefault("weight_per_meter", idx)
        if "wszystkie dolaczone pozycje" in t or "all items included" in t:
            header_map.setdefault("included", idx)
        if t == "strefa":
            header_map.setdefault("zone_col", idx)
        if t == "firma":
            header_map.setdefault("company_col", idx)
        if t == "element":
            header_map.setdefault("element_col", idx)
        if "nr elementu" in t:
            header_map.setdefault("element_nr_col", idx)
    return header_map


def fix_header_map_by_data(
    header_map: Dict[str, int],
    matrix: List[List[object]],
    first_data_row: int,
) -> Dict[str, int]:
    """
    Some files have swapped header labels (e.g. Nazwa/Name col contains integers=quantity,
    Ilość/Quantity col contains decimals=item weight). Detect by sampling data and fix.
    """
    sample = matrix[first_data_row : first_data_row + 5]
    if not sample:
        return header_map

    def col_vals(idx):
        return [row[idx] if idx < len(row) else "" for row in sample]

    def all_integer(idx):
        vals = [safe_float(v) for v in col_vals(idx) if str(v).strip()]
        return bool(vals) and all(v is not None and abs(v - round(v)) < 1e-9 for v in vals)

    def any_decimal(idx):
        vals = [safe_float(v) for v in col_vals(idx) if str(v).strip()]
        return any(v is not None and abs(v - round(v)) > 1e-9 for v in vals)

    fixed = dict(header_map)
    qty_col  = fixed.get("quantity")
    name_col = fixed.get("name")
    cls_col  = fixed.get("class")

    # Quantity must be a whole number. If the labeled quantity col has decimals and the
    # labeled name col has integers, the labels are swapped in this file.
    if qty_col is not None and any_decimal(qty_col):
        if name_col is not None and all_integer(name_col):
            fixed["quantity"] = name_col
            fixed["item_weight"] = qty_col   # old qty col = per-piece weight
            if cls_col is not None:
                fixed["name"] = cls_col
            else:
                fixed.pop("name", None)

    # If total_weight col is empty, fall back to weight_per_meter col (mislabeled in some files)
    tw_col  = fixed.get("total_weight")
    wpm_col = fixed.get("weight_per_meter")
    if wpm_col is not None:
        tw_empty = tw_col is None or not any(
            str(row[tw_col]).strip() for row in sample if tw_col < len(row)
        )
        if tw_empty:
            wpm_vals = [safe_float(v) for v in col_vals(wpm_col) if str(v).strip()]
            if wpm_vals and any(v is not None and v > 0 for v in wpm_vals):
                fixed["total_weight"] = wpm_col

    return fixed


def find_first_data_row(matrix: List[List[object]], header_row_idx: int, header_map: Dict[str, int]) -> int:
    num_col = header_map.get("number")
    if num_col is None:
        return header_row_idx + 1
    for r in range(header_row_idx + 1, len(matrix)):
        row = matrix[r]
        cell = row[num_col] if num_col < len(row) else ""
        if normalize_text(cell):
            return r
    return len(matrix)


def row_to_record(
    row: List[object],
    header_map: Dict[str, int],
    meta: Dict[str, str],
    source_row_number: int,
) -> Optional[Dict[str, object]]:
    def get(key: str) -> object:
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        return clean_value(row[idx])

    number = get("number")
    if not str(number).strip():
        zone_val    = str(get("zone_col")    or meta.get("zone", "")).strip()
        company_val = str(get("company_col") or meta.get("company", "")).strip()
        element_val = str(get("element_col")).strip()
        nr_val      = str(get("element_nr_col")).strip()
        if DEBUG:
            print(f"         [DEBUG row {source_row_number}] no 'number' col | zone={zone_val!r} company={company_val!r} element={element_val!r} nr={nr_val!r}")
        if element_val and nr_val:
            number = f"{zone_val}_{company_val}_{element_val}-{nr_val}"
            if DEBUG:
                print(f"         [DEBUG row {source_row_number}] reconstruido -> {number!r}")
    if not str(number).strip():
        return None

    rec = {
        "zone":              meta["zone"],
        "company":           meta["company"],
        "source_type":       meta["doc_type"],
        "revision":          meta.get("revision"),
        "number":            number,
        "name":              get("name"),
        "quantity":          get("quantity"),
        "length":            get("length"),
        "width":             get("width"),
        "thickness":         get("thickness"),
        "item_weight":       get("item_weight"),
        "total_weight":      get("total_weight"),
        "surface":           get("surface"),
        "class_grade":       get("class") or get("grade"),
        "included":          get("included"),
        "steel_type":        meta.get("steel_type", ""),
        "source_file":       meta["rel_path"],
        "source_row_number": source_row_number,
    }
    rec["number_key"] = normalize_key(rec["number"])
    return rec


def read_records_from_file(item: Dict[str, str]) -> List[Dict[str, object]]:
    path = item["path"]
    try:
        matrix = read_sheet_matrix(path)
    except Exception as exc:
        print(f"[ERROR:OPEN] {path} -> {exc}")
        return []

    header_row_idx = find_header_row(matrix)
    if header_row_idx is None:
        print(f"[WARN] Cabecera no encontrada en: {path}")
        return []

    header_map = build_header_map(matrix[header_row_idx])
    first_data_row = find_first_data_row(matrix, header_row_idx, header_map)
    header_map = fix_header_map_by_data(header_map, matrix, first_data_row)

    print(f"[HEADER] {item['rel_path']}")
    print(f"         header_row={header_row_idx + 1}, first_data_row={first_data_row + 1}, header_map={header_map}")
    if DEBUG:
        raw_header = [str(c) for c in matrix[header_row_idx]]
        print(f"         RAW HEADER: {raw_header}")

    records = []
    for r in range(first_data_row, len(matrix)):
        rec = row_to_record(matrix[r], header_map, item, r + 1)
        if rec is not None:
            records.append(rec)

    print(f"[ROWS_READ] {item['rel_path']} -> {len(records)}")
    return records


# =========================================================
# DISCOVERY
# =========================================================

def _find_steel_list_dir(section_dir: str) -> str:
    """Return the '4. List of steel elements…' subfolder directly under section_dir, or section_dir itself."""
    try:
        for entry in os.scandir(section_dir):
            if entry.is_dir() and "list of steel" in entry.name.lower():
                return entry.path
    except OSError:
        pass
    return section_dir


def collect_files(cfg: Dict) -> List[Dict[str, str]]:
    root_dir = cfg["root_dir"]
    found: List[Dict[str, str]] = []

    for top in cfg["search_folders"]:
        start_dir = os.path.join(root_dir, top)
        if not os.path.isdir(start_dir):
            print(f"[WARN] No existe: {start_dir}")
            continue

        walk_root = _find_steel_list_dir(start_dir)
        print(f"[SCAN] {walk_root}")

        for dirpath, dirnames, filenames in os.walk(walk_root):
            dirnames[:] = [d for d in dirnames if "archive" not in d.lower()]
            for filename in filenames:
                full_path = os.path.join(dirpath, filename)
                if not filename.lower().endswith(".xls"):
                    continue

                skip, reason = should_skip_file(full_path, cfg)
                if skip:
                    print(f"[SKIP:{reason}] {full_path}")
                    continue

                rel_path = os.path.relpath(full_path, root_dir)
                parts = Path(rel_path).parts
                company = detect_company(parts, filename)
                zone = get_top_area(rel_path)
                revision = extract_revision_from_text(rel_path)

                steel_type = ""
                for part in parts:
                    p_norm = normalize_text(part)
                    if "goracowalcowane" in p_norm or "hot-rolled" in p_norm or "hot rolled" in p_norm:
                        steel_type = "hot-rolled"
                        break
                    elif "zimnogiete" in p_norm or "cold-formed" in p_norm or "cold formed" in p_norm or "cold-rolled" in p_norm or "cold rolled" in p_norm:
                        steel_type = "cold-formed"
                        break

                item = {
                    "path":            full_path,
                    "rel_path":        rel_path,
                    "file_name":       filename,
                    "doc_type":        reason,
                    "company":         company,
                    "zone":            zone,
                    "revision":        revision,
                    "steel_type":      steel_type,
                }
                print(f"[USE:{reason}] {rel_path} | company={company}")
                found.append(item)

    return found


# =========================================================
# PARSING OUTPUT COLUMNS
# =========================================================

def parse_number_fields(
    number: object,
    zone_from_path: str = "",
    company_from_meta: str = "",
    revision_from_meta: object = None,
    prefix_pattern: str = "",
) -> Dict[str, object]:
    raw = str(number or "").strip()
    if not raw:
        rev_val = safe_int(revision_from_meta)
        return {
            "name_code":      "",
            "zone":           zone_from_path,
            "symbol":         company_from_meta if company_from_meta in {"GT", "KG"} else "",
            "element":        "",
            "element_number": "",
            "rev":            rev_val if rev_val is not None else "",
            "part1":          "",
            "part2":          "",
            "construction":   "",
        }

    s = raw.strip().replace("/", "_")
    s = re.sub(r"\s+", " ", s)
    if prefix_pattern:
        s = re.sub(prefix_pattern, "", s, flags=re.IGNORECASE)

    rev_val = extract_revision_from_text(s)
    if rev_val is None:
        rev_val = safe_int(revision_from_meta)
    s = re.sub(r"^R\s*\d+[-_ ]*", "", s, flags=re.IGNORECASE)

    zone = zone_from_path
    m_zone = re.match(r"^([A-Z]\d*[A-Z]?)[-_ ]+", s, flags=re.IGNORECASE)
    if m_zone:
        zone = m_zone.group(1).upper().replace(" ", "")
        s = s[m_zone.end():]

    symbol = company_from_meta if company_from_meta in {"GT", "KG"} else ""
    m_symbol = re.search(r"(?:^|[-_ ])(GT|KG)(?:[-_ ]|$)", s, flags=re.IGNORECASE)
    if m_symbol:
        symbol = m_symbol.group(1).upper()
        rest = s[m_symbol.end():].strip("_- ")
    else:
        rest = s.strip("_- ")

    m_num = re.search(r"(?:[-_])?(\d+)$", rest)
    element_number: object = ""
    element = ""
    if m_num:
        element_number = int(m_num.group(1))
        element = rest[:m_num.start()].strip("_- ")
    else:
        element = rest.strip("_- ")

    element = re.sub(r"\s+", " ", element).strip()
    if not element:
        element = "_"

    name_prefix = f"R{int(rev_val)}_" if rev_val not in (None, "") else ""

    if element and element_number != "":
        name_code = f"{name_prefix}{zone}_{symbol}_{element}-{element_number}".strip("_")
    elif element:
        name_code = f"{name_prefix}{zone}_{symbol}_{element}".strip("_")
    elif element_number != "":
        name_code = f"{name_prefix}{zone}_{symbol}-{element_number}".strip("_")
    else:
        name_code = raw

    construction = ""
    if isinstance(element_number, int):
        n = element_number
        if 0 <= n < 500:
            construction = "wall"
        elif (500 <= n <= 599) or (700 <= n <= 799) or (900 <= n <= 999):
            construction = "ceiling sub"
        elif (600 <= n <= 699) or (800 <= n <= 899):
            construction = "panel sub"

    return {
        "name_code":      name_code,
        "zone":           zone,
        "symbol":         symbol,
        "element":        element,
        "element_number": element_number,
        "rev":            int(rev_val) if rev_val not in (None, "") else "",
        "part1":          "",
        "part2":          "",
        "construction":   construction,
    }


def count_items(value: object) -> int:
    s = str(value or "").strip().rstrip(",").strip()
    if not s:
        return 0
    return s.count(",") + 1


WELDED_ELEMENTS = {"stp", "k", "s", "b", "ks"}


def weld_allowance_rate(steel_type: str, included: str, element: str) -> float:
    """
    1.8% weld allowance applies when:
    - hot-rolled AND (included > 1 OR element in WELDED_ELEMENTS when included unknown)
    Cold-formed never gets the allowance.
    """
    t = steel_type.lower()
    if "cold" in t or "zimno" in t:
        return 0.0
    s = str(included or "").strip().rstrip(",").strip()
    if s:
        return 0.018 if s.count(",") + 1 > 1 else 0.0
    return 0.018 if element.strip().lower() in WELDED_ELEMENTS else 0.0


def build_output_row(rec: Dict[str, object], prefix_pattern: str = "") -> Dict[str, object]:
    parsed = parse_number_fields(
        rec.get("number", ""),
        str(rec.get("zone", "")),
        str(rec.get("company", "")),
        rec.get("revision"),
        prefix_pattern=prefix_pattern,
    )

    rate = weld_allowance_rate(
        rec.get("steel_type", ""),
        rec.get("included", ""),
        parsed["element"],
    )

    # Source values (raw from file)
    item_weight_source  = excel_number(rec.get("item_weight", ""))
    total_weight_source = excel_number(rec.get("total_weight", ""))

    # Recalculate item weight from total/quantity for better precision
    qty_val = safe_float(rec.get("quantity", ""))
    tw_val  = safe_float(rec.get("total_weight", ""))
    if qty_val and qty_val > 0 and tw_val is not None:
        item_weight_calc = round(tw_val / qty_val, 6)
    else:
        item_weight_calc = item_weight_source

    gross_weight_source = round(tw_val * (1 + rate), 3) if tw_val is not None else ""

    return {
        "name_code":            rec.get("number", ""),
        "zone":                 parsed["zone"],
        "symbol":               parsed["symbol"],
        "element":              parsed["element"],
        "element_number":       parsed["element_number"],
        "rev":                  parsed["rev"],
        "construction":         parsed["construction"],
        "steel_type":           rec.get("steel_type", "") if parsed["construction"] == "ceiling sub" else "",
        "quantity":             excel_number(rec.get("quantity", "")),
        "name":                 rec.get("name", ""),
        "length":               excel_number(rec.get("length", "")),
        "item_weight":          item_weight_calc,
        "weld_pct":             round(rate * 100, 1),
        "item_weight_source":   item_weight_source,
        "total_weight_source":  total_weight_source,
        "gross_weight_source":  excel_number(gross_weight_source),
        "surface":             excel_number(rec.get("surface", "")),
        "items_count":         count_items(rec.get("included", "")),
        "included":            rec.get("included", ""),
        "source_file":         rec.get("source_file", ""),
        "source_row_number":   excel_number(rec.get("source_row_number", "")),
    }


# =========================================================
# OUTPUT
# =========================================================

def write_output_xlsx(output_path: str, rows: List[Dict[str, object]]) -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "assembled"

    # Number formats keyed by 1-based column index
    col_fmts = {
        5:  "0",       # element_number
        6:  "0",       # rev
        9:  "0",       # quantity
        11: "0.###",   # length
        12: "0.####",  # item_weight (recalculated)
        13: "0.00",    # Weight netto (kg) — formula
        14: "0.00",    # Weight brutto (kg) — formula
        15: "0.0",     # weld_pct
        16: "0.####",  # item_weight_source
        17: "0.00",    # total_weight_source
        18: "0.00",    # gross_weight_source
        19: "0.000",   # Netto check (ratio, empty when OK)
        20: "0.###",   # surface
        21: "0",       # items_count
        24: "0",       # source_row_number
    }

    ws.append(OUTPUT_COLUMNS)

    qty_col    = get_column_letter(OUTPUT_COLUMNS.index("Quantity") + 1)
    iw_col     = get_column_letter(OUTPUT_COLUMNS.index("Weight netto (kg/pcs.)") + 1)
    netto_col  = get_column_letter(OUTPUT_COLUMNS.index("Weight netto (kg)") + 1)
    weld_col   = get_column_letter(OUTPUT_COLUMNS.index("Weld allowance (%)") + 1)
    iw_src_col = get_column_letter(OUTPUT_COLUMNS.index("Weight netto (kg/pcs.) source") + 1)
    tw_src_col = get_column_letter(OUTPUT_COLUMNS.index("Weight netto (kg) source") + 1)

    for row in rows:
        cur = ws.max_row + 1
        values = [
            row.get("name_code", ""),                                       # 1
            row.get("zone", ""),                                            # 2
            row.get("symbol", ""),                                          # 3
            row.get("element", ""),                                         # 4
            row.get("element_number", ""),                                  # 5
            row.get("rev", ""),                                             # 6
            row.get("construction", ""),                                    # 7
            row.get("steel_type", ""),                                      # 8
            row.get("quantity", ""),                                        # 9
            row.get("name", ""),                                            # 10
            row.get("length", ""),                                          # 11
            row.get("item_weight", ""),                                     # 12  recalculated
            f"=ROUND({qty_col}{cur}*{iw_col}{cur},2)",                      # 13  Weight netto (kg)
            f"=ROUND({netto_col}{cur}*(1+{weld_col}{cur}/100),2)",          # 14  Weight brutto (kg)
            row.get("weld_pct", 0.0),                                       # 15
            row.get("item_weight_source", ""),                              # 16  source
            row.get("total_weight_source", ""),                             # 17  source
            row.get("gross_weight_source", ""),                             # 18  source
            f'=IF({tw_src_col}{cur}="","",IF(ABS({qty_col}{cur}*{iw_src_col}{cur}-{tw_src_col}{cur})/MAX({tw_src_col}{cur},0.001)>0.05,{qty_col}{cur}*{iw_src_col}{cur}/{tw_src_col}{cur},""))',  # 19
            row.get("surface", ""),                                         # 20
            row.get("items_count", 0),                                      # 21
            row.get("included", ""),                                        # 22
            row.get("source_file", ""),                                     # 23
            row.get("source_row_number", ""),                               # 24
        ]
        ws.append(values)
        for col_idx, fmt in col_fmts.items():
            ws.cell(row=cur, column=col_idx).number_format = fmt

    n_data   = len(rows)
    n_cols   = len(OUTPUT_COLUMNS)
    last_col = get_column_letter(n_cols)   # "T"
    tot_row  = n_data + 2                  # header + data + totals

    # Weight brutto column (1-based → letter)
    gross_idx = OUTPUT_COLUMNS.index("Weight brutto (kg)") + 1
    gross_col = get_column_letter(gross_idx)

    # Write SUBTOTAL formula in the totals row cell
    cell = ws[f"{gross_col}{tot_row}"]
    cell.value = "=SUBTOTAL(109,[Weight brutto (kg)])"
    cell.number_format = "0.###"

    # Build table columns — mark gross weight column as custom total
    tab_cols = []
    for i, name in enumerate(OUTPUT_COLUMNS, 1):
        tc = TableColumn(id=i, name=name)
        if i == gross_idx:
            tc.totalsRowFunction = "custom"
            tc.totalsRowFormula  = TableFormula("SUBTOTAL(109,[Weight brutto (kg)])")
        tab_cols.append(tc)

    # Create table spanning header + data + totals row
    tab = Table(displayName="SteelCatalog", ref=f"A1:{last_col}{tot_row}")
    tab.totalsRowCount = 1
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    for tc in tab_cols:
        tab.tableColumns.append(tc)
    ws.add_table(tab)

    widths = [28, 10, 20, 14, 16, 12, 18, 20, 10, 32, 14, 18, 18, 20, 14, 20, 18, 18, 12, 22, 12, 34, 60, 14]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    wb.save(output_path)


# =========================================================
# MAIN
# =========================================================

def run_project(name: str, cfg: Dict) -> None:
    root_dir = cfg["root_dir"]
    output_path = os.path.join(OUT_DIR, cfg["output_name"])
    prefix_pattern = cfg["prefix_pattern"]

    print(f"\n{'=' * 60}")
    print(f"Proyecto: {name}  |  {root_dir}")
    print(f"{'=' * 60}")

    if not os.path.isdir(root_dir):
        print(f"[ERROR] Ruta no existe o no accesible: {root_dir}")
        return

    files = collect_files(cfg)
    if not files:
        print("[ERROR] No se han encontrado archivos válidos.")
        return

    print(f"Archivos válidos: {len(files)}")

    final_by_number: Dict[str, Dict[str, object]] = {}
    skipped_out_of_range = 0

    for item in files:
        records = read_records_from_file(item)
        for rec in records:
            parsed = parse_number_fields(
                rec.get("number", ""),
                str(rec.get("zone", "")),
                str(rec.get("company", "")),
                rec.get("revision"),
                prefix_pattern=prefix_pattern,
            )
            en = parsed["element_number"]
            if not isinstance(en, int) or en < 0 or en > 999:
                skipped_out_of_range += 1
                continue
            key = str(rec["number_key"])
            if key not in final_by_number:
                final_by_number[key] = rec

    print(f"Elementos fuera de rango (>=1000 o sin número): {skipped_out_of_range}")

    final_rows = [build_output_row(rec, prefix_pattern=prefix_pattern) for rec in final_by_number.values()]
    final_rows.sort(key=lambda x: (str(x.get("zone", "")), str(x.get("symbol", "")), str(x.get("name_code", ""))))

    print(f"Elementos finales: {len(final_rows)}")
    write_output_xlsx(output_path, final_rows)
    print(f"OK -> {output_path}")


def ask_project() -> Dict:
    options = list(PROJECT_CONFIGS.keys())
    print("\n¿Qué proyecto quieres generar?")
    for i, name in enumerate(options, start=1):
        print(f"  {i}. {name}")
    print(f"  {len(options) + 1}. Todos")
    print()

    while True:
        raw = input("Selecciona una opción: ").strip()
        if raw.upper() in PROJECT_CONFIGS:
            return {raw.upper(): PROJECT_CONFIGS[raw.upper()]}
        try:
            n = int(raw)
            if 1 <= n <= len(options):
                key = options[n - 1]
                return {key: PROJECT_CONFIGS[key]}
            if n == len(options) + 1:
                return PROJECT_CONFIGS
        except ValueError:
            pass
        print(f"  Opción no válida. Introduce un número del 1 al {len(options) + 1}.")


def main() -> None:
    args = [a.upper() for a in sys.argv[1:]]

    if args:
        unknown = [a for a in args if a not in PROJECT_CONFIGS]
        if unknown:
            print(f"[ERROR] Proyecto(s) desconocido(s): {unknown}. Opciones: {list(PROJECT_CONFIGS.keys())}")
            sys.exit(1)
        selected = {k: PROJECT_CONFIGS[k] for k in args}
    else:
        selected = ask_project()

    for name, cfg in selected.items():
        run_project(name, cfg)


if __name__ == "__main__":
    main()
